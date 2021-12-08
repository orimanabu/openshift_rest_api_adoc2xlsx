#!/usr/bin/env python3

# https://docs.openshift.com/container-platform/4.8/rest_api/index.html

import os
import re
import csv
import sys
import html
import json
import logging
import datetime
import argparse
import openpyxl
import subprocess

url_prefix = 'https://docs.openshift.com/container-platform'
ocp_version = ''

def get_ocp_version_from_dotgit(repodir):
    with open('/'.join([repodir, '.git', 'HEAD']), 'r') as f:
        line = f.readline().rstrip()
        m = re.search(r'^.*enterprise-([0-9]+\.[0-9]+)$', line)
        return m.group(1)


def adoc_path2url(path):
    dirname, adoc = os.path.split(path)
    dirname, category = os.path.split(dirname)
    topdir, restapi = os.path.split(dirname)
    version = get_ocp_version_from_dotgit(topdir)
    return version, '/'.join([url_prefix, version, restapi, category, adoc.replace('.adoc', '.html')])

def xref2url(path, version):
    return '/'.join([url_prefix, version, 'rest_api', path[3:].replace('.adoc', '.html')])

def skip_header(file):
    for line in file:
        if line.startswith('= '):
            m = re.search(r'^= (.*) \[.*$', line)
            title = m.group(1)
        if line.startswith('== API endpoints'):
            logger.info('** API endoints')
            break
    return title

def parse_table(file):
    for line in file:
        if line.startswith('[cols='):
            continue
        if line.startswith('|==='):
            line = file.readline() # for table header
            break;

    table_str = ''
    for line in file:
        if line.startswith('|==='):
            break;
        table_str = table_str + html.unescape(line)

    array = re.split(r'\|[ \n]', table_str)
    array.pop(0) # discard empty
    logger.info('*** array: ' + str(array))
    return array

def parse_global_path_parameters(file):
    mode = 'Global path parameters'
    apiref[mode] = []
    array = parse_table(file)

    while True:
        if len(array) == 0:
            break
        param = {}
        x = array.pop(0)
        param['Parameter'] = x.rstrip()[1:-1]

        x = array.pop(0)
        param['Type'] = x.rstrip()[1:-1]

        x = array.pop(0)
        param['Description'] = x

        apiref[mode].append(param)

def parse_global_query_parameters(file):
    mode = 'Global query parameters'
    apiref[mode] = []
    array = parse_table(file)

    while True:
        if len(array) == 0:
            break
        param = {}
        x = array.pop(0)
        param['Parameter'] = x.rstrip()[1:-1]

        x = array.pop(0)
        param['Type'] = x.rstrip()[1:-1]

        x = array.pop(0)
        param['Description'] = x

        apiref[mode].append(param)

def is_http_method(line):
    http_methods = ['GET', 'PUT', 'POST', 'DELETE', 'PATCH'];
    for method in http_methods:
        if line.lstrip().startswith('`' + method + '`'):
            return True, method
    return False, ""

def parse_http_method_xref(line):
    if line.startswith('xref:'):
        m = re.search(r'xref:(.*)\[`([^]]+)`\]', line)
        path = m.group(1)
        value = m.group(2)
        hyperlink = xref2url(path, ocp_version)
        logger.info('*** ocp_version: ' + ocp_version)
        return value, hyperlink
    return line[1:-1], None

def parse_http_method(file):
    mode = 'HTTP method'
    apiref[mode] = []

    param = {}
    for line in file:
        logger.info('*** line: ' + line.rstrip())
        b, method = is_http_method(line)
        if b:
            logger.info('*** HTTP Method: {}'.format(method))
            if param.get('Method'):
                logger.info('*** append(): ' + json.dumps(param))
                apiref[mode].append(param)
                param = {}
            param['Method'] = method
        elif line.startswith('\n'):
            # print('*** \\n')
            pass
        elif line.startswith('Description::'):
            line = file.readline()
            desc = line.rstrip().lstrip()
            logger.info('*** desc: {}'.format(desc))
            param['Description'] = desc
        elif line.startswith('.Query parameters') or line.startswith('.Body parameters'):
            section = line[1:-1]
            array = parse_table(file)
            logger.info('*** array: ' + json.dumps(array))
            param[section] = []
            while True:
                if len(array) == 0:
                    break
                tmp = {}
                tmp['Parameter'] = array.pop(0).rstrip()[1:-1]
                value, hyperlink = parse_http_method_xref(array.pop(0).rstrip())
                logger.info('*** parameter type: value={}, link={}'.format(value, hyperlink))
                tmp['Type'] = {}
                tmp['Type']['value'] = value
                tmp['Type']['hyperlink'] = hyperlink
                tmp['Description'] = array.pop(0)
                param[section].append(tmp)
        elif line.startswith('.HTTP responses'):
            section = line[1:-1]
            array = parse_table(file)
            logger.info('*** array: ' + json.dumps(array))
            param[section] = []
            while True:
                if len(array) == 0:
                    break
                tmp = {}
                tmp['HTTP code'] = array.pop(0).rstrip()
                value, hyperlink = parse_http_method_xref(array.pop(0).rstrip())
                logger.info('*** response body: value={}, link={}'.format(value, hyperlink))
                tmp['Response body'] = {}
                tmp['Response body']['value'] = value
                tmp['Response body']['hyperlink'] = hyperlink
                param[section].append(tmp)
        elif line.startswith('=== /api'):
            logger.info('*** break: ' + line)
            break

    logger.info('*** append(): ' + json.dumps(param))
    apiref[mode].append(param)
    return line

def csv_indent_section(row, section):
    if section == 'Global path parameters':
        pass
    elif section == 'Global query parameters':
        row.append('') # Gpp Parameter
        row.append('') # Gpp Type
        row.append('') # Gpp Description
    elif section == 'HTTP method':
        row.append('') # Gpp Parameter
        row.append('') # Gpp Type
        row.append('') # Gpp Description
        row.append('') # Gqp Parameter
        row.append('') # Gqp Type
        row.append('') # Gqp Description

def csv_indent_subsection(row, subsection):
    if subsection == 'Query parameters':
        pass
    elif subsection == 'Body parameters':
        row.append('') # Qp Parameter
        row.append('') # Qp Type
        row.append('') # Qp Description
    elif subsection == 'HTTP responses':
        row.append('') # Qp Parameter
        row.append('') # Qp Type
        row.append('') # Qp Description
        row.append('') # Bp Parameter
        row.append('') # Bp Type
        row.append('') # Bp Description

def add_section_global_params(row, rows, section, apiref):
    if apiref.get(section):
        for item in apiref[section]:
            row.append(item['Parameter'])
            row.append(item['Type'])
            row.append(item['Description'])
            rows.append(row)
            row = []

def add_section_http_method(row, rows, section, subsection, apiref):
    if not apiref.get(section):
        return
    for method in apiref[section]:
        row.append(method['Method'])
        row.append(subsection)

        if subsection == 'Query parameters' or subsection == 'Body parameters':
            array = method.get(subsection)
            if array:
                for item in array:
                    row.append(item['Parameter'])
                    row.append(item['Type'])
                    row.append(item['Description'])

        elif subsection == 'HTTP responses':
            array = method.get(subsection)
            if array:
                for item in array:
                    row.append(item['HTTP code'])
                    row.append(item['Response body'])

        rows.append(row)
        row = []

def build_csv(allapiref):
    rows = []
    rows.append([allapiref['url']])
    rows.append([
        'Endpoint',
        'Section',
        'Subsection',
        'HTTP method',
        'Global path parameters - Parameter',
        'Global path parameters - Type',
        'Global path parameters - Description',
        'Global query parameters - Parameter',
        'Global query parameters - Type',
        'Global query parameters - Description',
        'HTTP method - Query parameters - Parameter',
        'HTTP method - Query parameters - Type',
        'HTTP method - Query parameters - Description',
        'HTTP method - Body parameters - Parameter',
        'HTTP method - Body parameters - Type',
        'HTTP method - Body parameters - Description',
        'HTTP method - HTTP responses - HTTP code',
        'HTTP method - HTTP responses - HTTP Response body',
    ])
    for apiref in allapiref['items']:
        row = []

        sections = ['Global path parameters', 'Global query parameters', 'HTTP method']
        for section in sections:
            if section == 'Global path parameters' or section == 'Global query parameters':
                if apiref.get(section):
                    for item in apiref[section]:
                        row.append(apiref['Endpoint'])
                        row.append(section)
                        row.append('') # subsection
                        row.append('') # http method
                        csv_indent_section(row, section)
                        row.append(item['Parameter'])
                        row.append(item['Type'])
                        row.append(item['Description'])
                        rows.append(row[:])
                        row = []
            else:
                subsections = ['Query parameters', 'Body parameters', 'HTTP responses']
                if apiref.get(section):
                    for method in apiref[section]:
                        for subsection in subsections:
                            if method.get(subsection):
                                for item in method[subsection]:
                                    if subsection == 'Query parameters' or subsection == 'Body parameters':
                                        row.append(apiref['Endpoint'])
                                        row.append(section)
                                        row.append(subsection)
                                        row.append(method['Method'])
                                        csv_indent_section(row, section)
                                        csv_indent_subsection(row, subsection)
                                        row.append(item['Parameter'])
                                        row.append(item['Type'])
                                        row.append(item['Description'])
                                        rows.append(row[:])
                                        row = []
                                    elif subsection == 'HTTP responses':
                                        row.append(apiref['Endpoint'])
                                        row.append(section)
                                        row.append(subsection)
                                        row.append(method['Method'])
                                        csv_indent_section(row, section)
                                        csv_indent_subsection(row, subsection)
                                        row.append(item['HTTP code'])
                                        row.append(item['Response body'])
                                        rows.append(row[:])
                                        row = []
    return rows

def print_csv(allapiref, filename):
    rows = build_csv(allapiref)

    if filename == '-' or filename == None:
        f = sys.stdout
    else:
        f = open(filename, 'w', newline='')

    writer = csv.writer(f, quoting=csv.QUOTE_ALL)
    writer.writerows(rows)

def color_cell(sheet, row, col, col_max, color):
    for i in range(col, col_max + 1):
        sheet.cell(row, i).fill = color

def print_xlsx(title, allapiref, filename):
    fill_endpoint = openpyxl.styles.PatternFill(patternType='solid', fgColor='D9EAD3')
    fill_section = openpyxl.styles.PatternFill(patternType='solid', fgColor='FCE5CD')
    fill_method = openpyxl.styles.PatternFill(patternType='solid', fgColor='CFE2F3')
    fill_http = openpyxl.styles.PatternFill(patternType='solid', fgColor='FFF2CC')

    book = openpyxl.Workbook()
    # sheet = book.create_sheet(title)
    sheet = book.active
    sheet.title = 'Summary'
    y = 1
    for ep in allapiref['summary'].keys():
        sheet.cell(y, 1, ep)
        color_cell(sheet, y, 1, 3, fill_endpoint)
        y = y + 1
        for item in allapiref['summary'][ep]:
            # sheet.cell(y, 2, item)
            sheet.append(['', item['method'], item['description']])
            y = y + 1
    sheet.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 90

    sheet = book.create_sheet(title)

    x = 1
    y = 1
    max_fill_x = 6
    cell = sheet.cell(y, x, allapiref['url'])
    cell.hyperlink = allapiref['url']
    cell.style = "Hyperlink"
    # color_cell(sheet, y, x, max_fill_x, fill_endpoint)
    y = y + 1

    for apiref in allapiref['items']:
        sheet.cell(y, x, apiref['Endpoint'])
        color_cell(sheet, y, 1, max_fill_x, fill_endpoint)
        y = y + 1

        for section in ['Global path parameters', 'Global query parameters']:
            sheet.append(['', section, '', 'Parameter', 'Type', 'Description']) # header
            color_cell(sheet, y, 2, max_fill_x, fill_section)
            y = y + 1
            if not apiref.get(section):
                sheet.append(['', '', '', '-', '-', '-'])
                y = y + 1
            else:
                for item in apiref[section]:
                    sheet.append(['', '', '', item['Parameter'], item['Type'], item['Description'].rstrip()])
                    y = y + 1

        section = 'HTTP method'
        sheet.append(['', section, 'Method', '', '', 'Description']) # header
        color_cell(sheet, y, 2, max_fill_x, fill_section)
        y = y + 1

        for method in apiref[section]:
            sheet.append(['', '', method['Method'], '', '', method['Description']])
            color_cell(sheet, y, 3, max_fill_x, fill_method)
            y = y + 1

            for subsection in ['Query parameters','Body parameters']:
                sheet.append(['', # Endpoint
                              '', # Section
                              '({}: {})'.format(method['Method'], subsection), # Method + Subsection
                              'Parameter',
                              'Type',
                              'Description'])
                color_cell(sheet, y, 3, max_fill_x, fill_http)
                y = y + 1
                if not method.get(subsection):
                    sheet.append(['', '', '', '-', '-', '-'])
                    y = y + 1
                else:
                    for item in method[subsection]:
                        # sheet.append(['', '', '', item['Parameter'], item['Type'], item['Description']])
                        sheet.cell(y, 4, item['Parameter'])
                        t = item['Type']
                        cell = sheet.cell(y, 5, t.get('value'))
                        if t.get('hyperlink'):
                            cell.hyperlink = t.get('hyperlink')
                            cell.style = "Hyperlink"
                        sheet.cell(y, 6, item['Description'].rstrip())
                        y = y + 1

            subsection = 'HTTP responses'
            sheet.append(['', # Endpoint
                          '', # Section
                          '({}: {})'.format(method['Method'], subsection), # Method + Subsection
                          'HTTP code',
                          'Response body'])
            color_cell(sheet, y, 3, max_fill_x, fill_http)
            y = y + 1
            if not method.get(subsection):
                sheet.append(['', '', '', '-', '-'])
                y = y + 1
            else:
                for item in method[subsection]:
                    # sheet.append(['', '', '', item['HTTP code'], item['Response body']])
                    sheet.cell(y, 4, item['HTTP code'])
                    rb = item['Response body']
                    cell = sheet.cell(y, 5, rb.get('value'))
                    if rb.get('hyperlink'):
                        cell.hyperlink = rb.get('hyperlink')
                        cell.style = "Hyperlink"
                    y = y + 1

        sheet.append([''])
        y = y + 1

    sheet.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 25
    sheet.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 30
    sheet.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 30
    sheet.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 20
    sheet.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 90
    for col in sheet.iter_cols(min_col=6, max_col=6):
        for cell in col:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    sheet = book.create_sheet('Info')
    commit_sha = subprocess.run("git show --format=oneline --no-patch | awk '{print $1'}", shell=True, capture_output=True, encoding='utf-8').stdout
    sheet.cell(1, 1, 'This book is generated by adoc2xlsx.py at {}.'.format(datetime.datetime.today()))
    sheet.cell(2, 1, 'The commit id of adoc2xlsx.py is: {}.'.format(commit_sha.rstrip()))
    cell = sheet.cell(3, 1, 'https://github.com/orimanabu/openshift_rest_api_adoc2xlsx.git')
    cell.hyperlink = 'https://github.com/orimanabu/openshift_rest_api_adoc2xlsx.git'
    cell.style = "Hyperlink"

    if filename == None:
        print("Error: needs '--output filename' when output format is xlsx.")
        exit(1)
    book.save(filename)

def print_json(allapiref, filename):
    if filename == '-' or filename == None:
        print(json.dumps(allapiref))
    else:
        with open(filename, 'w') as f:
            f.write(json.dumps(allapiref))


allapiref = {}
if __name__ == '__main__':
    # print(sys.argv)

    parser = argparse.ArgumentParser()
    parser.add_argument('adoc', help='adoc path')
    parser.add_argument('-f', '--format', default='json', choices=['json', 'csv', 'xlsx'], help='output format')
    parser.add_argument('-o', '--output', help='output file name')
    parser.add_argument('-d', '--debug', action='store_true')
    args = parser.parse_args()

    sh = logging.StreamHandler(stream=sys.stdout)
    # sh.setLevel(logging.WARNING)
    logger = logging.getLogger('xxx')
    if args.debug:
        logger.setLevel(logging.INFO)
    logger.addHandler(sh)

    file = open(args.adoc, 'r')
    ocp_version, allapiref['url'] = adoc_path2url(args.adoc)
    logger.info("** ocp_version: " + ocp_version)
    allapiref['items'] = []

    title = ''

    apiref = {}
    endpoint = ''
    mode = ''
    prev = ''
    summary = False
    for line in file:
        logger.info('?? ' + line.rstrip())
        if prev:
            line = prev
            prev = ''
        if line.startswith('= '):
            m = re.search(r'^= (.*) \[.*$', line)
            title = m.group(1)
        elif line.startswith('== API endpoints'):
            summary = True
            allapiref['summary'] = {}
            logger.info('** summary start')
        elif summary and line.startswith('* '):
            m = re.search(r'^\* `(.*)`', line)
            endpoint = m.group(1)
            allapiref['summary'][endpoint] = []
            logger.info('** summary endpoint: ' + endpoint)
        elif summary and line.startswith('- '):
            m = re.search(r'^- `(.*)`: (.*)', line.rstrip())
            method = m.group(1)
            desc = m.group(2)
            allapiref['summary'][endpoint].append({'method': method, 'description': desc})
            logger.info('** summary method: ' + method)
        elif line.startswith('=== /api'):
            summary = False
            m = re.search(r'^=== (.*)$', line.rstrip())
            endpoint = m.group(1)
            logger.info('** endpoint: {}'.format(endpoint))
            if apiref.get('Endpoint'):
                logger.info("** allapiref['items'].append(): {}".format(apiref.get('Endpoint')))
                allapiref['items'].append(apiref)
                apiref = {}
            apiref['Endpoint'] = endpoint
        elif line.startswith('.Global path parameters'):
            parse_global_path_parameters(file)
        elif line.startswith('.Global query parameters'):
            parse_global_query_parameters(file)
        elif line.startswith('HTTP method::'):
            prev = parse_http_method(file)


    logger.info("** allapiref['items].append(): {}".format(apiref.get('Endpoint')))
    allapiref['items'].append(apiref)

    if args.format == 'csv':
        print_csv(allapiref, args.output)
    elif args.format == 'xlsx':
        print_xlsx(title, allapiref, args.output)
    else:
        print_json(allapiref, args.output)
